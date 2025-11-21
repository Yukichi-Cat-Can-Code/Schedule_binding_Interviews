"""
API Views for Interview Scheduler - MongoDB compatible version
"""
from rest_framework import status
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response
from rest_framework.views import APIView
from django.conf import settings
import pandas as pd
from datetime import datetime
from bson import ObjectId
import time

from api.mongo_models import (
    Position, InterviewSession, Company,
    Applicant, Interviewer, Room, Schedule,
    AlgorithmConfig, ScheduleResult, ActionLog
)
from scheduler.genetic_algorithm import GeneticAlgorithm
from scheduler.genetic_algorithm_variant import GeneticAlgorithmVariant
from scheduler.genetic_algorithm_variant2 import GeneticAlgorithmVariant2
from scheduler.genetic_algorithm_variant3 import GeneticAlgorithmVariant3
import random  # Needed for top-k selection and variant cycling
from api.auth_utils import User, derive_company_id, get_request_user

# Utilities
def to_json_safe(value):
    """Recursively convert ObjectId and datetime to JSON-safe primitives."""
    from bson import ObjectId
    from datetime import datetime as _dt

    if isinstance(value, dict):
        return {k: to_json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [to_json_safe(v) for v in value]
    if isinstance(value, tuple):
        return tuple(to_json_safe(v) for v in value)
    if isinstance(value, ObjectId):
        return str(value)
    if isinstance(value, _dt):
        return value.isoformat()
    return value


def normalize_data(items):
    """Add 'id' alias for '_id' to make data compatible with schedulers"""
    for item in items:
        if '_id' in item:
            item['id'] = item['_id']
    return items


@api_view(["GET"])
def whoami(request):
    """Debug endpoint to inspect current authenticated user and company.

    Helps verify token -> company_id mapping in multi-tenant scenarios.
    """
    user = get_request_user(request)
    company_id = derive_company_id(request)

    if not user:
        return Response(
            {"authenticated": False, "company_id": company_id},
            status=status.HTTP_200_OK,
        )

    return Response(
        {
            "authenticated": True,
            "user": {
                "id": str(user.get("_id")),
                "username": user.get("username"),
                "role": user.get("role"),
                "company_id": user.get("company_id"),
            },
            "company_id": company_id,
        },
        status=status.HTTP_200_OK,
    )


def log_action(request, action_type: str, **kwargs):
    """Create an action log entry.

    Lightweight helper; failures should not break main flow.
    """
    try:
        user = getattr(request, "user", None)
        user_id = None
        user_email = None
        role = None
        if isinstance(user, User):
            user_id = str(user.id) if getattr(user, "id", None) is not None else None
            user_email = getattr(user, "email", None)
            role = getattr(user, "role", None)

        company_id = kwargs.pop("company_id", None) or derive_company_id(request)

        doc = {
            "action_type": action_type,
            "user_id": user_id,
            "user_email": user_email,
            "role": role,
            "company_id": company_id,
            "created_at": datetime.now(),
        }
        doc.update(kwargs)
        ActionLog.create(doc)
    except Exception:
        # Logging must never break the main request
        return


def _get_active_session_or_none(company_id: str | None = None):
    """Return active session optionally scoped by company."""
    try:
        if company_id:
            sessions = InterviewSession.find_all({'is_active': True, 'company_id': company_id}, limit=1, sort=[('created_at', -1)])
        else:
            sessions = InterviewSession.find_all({'is_active': True}, limit=1, sort=[('created_at', -1)])
        return sessions[0] if sessions else None
    except Exception:
        return None


def _parse_iso(dt_str):
    try:
        return datetime.fromisoformat(dt_str)
    except Exception:
        return None


def _overlaps(a_start, a_end, b_start, b_end) -> bool:
    return a_start < b_end and b_start < a_end


def _filter_by_session_constraints(schedule_data, current_session):
    """Apply session constraints before saving:
    - No overlap with schedules of other sessions that time-overlap current session window
    - Interviewer cannot work more than 8h continuously per day in current session
    Returns filtered list and stats.
    """
    if not current_session:
        return schedule_data, {'skipped_conflicts': 0, 'skipped_overtime': 0}

    # Load all schedules to check cross-session conflicts
    existing = Schedule.find_all()
    cur_id = current_session.get('_id')
    start_win = _parse_iso(current_session.get('start_date'))
    end_win = _parse_iso(current_session.get('end_date'))

    other = []
    for s in existing:
        if s.get('session_id') == cur_id:
            continue
        s_start = _parse_iso(s.get('start_time'))
        s_end = _parse_iso(s.get('end_time'))
        if not s_start or not s_end:
            continue
        # keep only those that overlap window
        if start_win and end_win and _overlaps(s_start, s_end, start_win, end_win):
            other.append(s)

    # track per-interviewer continuous minutes per day
    from collections import defaultdict
    work_minutes = defaultdict(lambda: defaultdict(int))  # interviewer_id -> date -> minutes

    filtered = []
    skipped_conflicts = 0
    skipped_overtime = 0

    for entry in schedule_data:
        ns = _parse_iso(entry.get('start_time'))
        ne = _parse_iso(entry.get('end_time'))
        if not ns or not ne:
            continue

        # Cross-session conflicts (interviewer/room)
        conflict = False
        for s in other:
            if (s.get('interviewer_id') == entry.get('interviewer_id') or
                s.get('room_id') == entry.get('room_id')):
                os = _parse_iso(s.get('start_time'))
                oe = _parse_iso(s.get('end_time'))
                if os and oe and _overlaps(ns, ne, os, oe):
                    conflict = True
                    break
        if conflict:
            skipped_conflicts += 1
            continue

        # 8-hour/day interviewer rule inside current session
        key_day = ns.date().isoformat()
        iv = entry.get('interviewer_id')
        minutes = int((ne - ns).total_seconds() // 60)
        if work_minutes[iv][key_day] + minutes > 8 * 60:
            skipped_overtime += 1
            continue
        work_minutes[iv][key_day] += minutes
        filtered.append(entry)

    return filtered, {'skipped_conflicts': skipped_conflicts, 'skipped_overtime': skipped_overtime}


@api_view(['GET'])
def current_company(request):
    """Return the current company for the authenticated user.

    This endpoint is intentionally strict:
    - If the token cannot be resolved to a user/company_id -> 404
    - If the referenced company document does not exist -> 404

    We avoid auto-creating fallback companies here because that hides
    data problems (e.g. seeded Demo Company vs. empty "Company for demo").
    The seed script `scripts/import_sample_data.py` is the source of truth
    for attaching users (like "demo") to real companies with data.
    """
    try:
        company_id = derive_company_id(request)
        fallback_used = False
        if not company_id:
            # Fallback: allow client to provide company id (useful when Authorization
            # header is not forwarded by proxies). The frontend stores `auth_company_id`
            # in localStorage after login; include it as `X-Auth-Company-Id` header or
            # as query param `company_id` when calling `/companies/current/`.
            company_id = request.headers.get('X-Auth-Company-Id') or request.GET.get('company_id')
            if company_id:
                fallback_used = True

        if not company_id:
            return Response(
                {'error': 'No company associated with user'},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Company id may be an ObjectId string or a company code; try both
        company = None
        try:
            company = Company.find_one({'_id': ObjectId(company_id)})
        except Exception:
            # not a valid ObjectId, try lookup by code
            company = Company.find_one({'code': company_id})

        if not company:
            return Response(
                {'error': 'Company not found for this user'},
                status=status.HTTP_404_NOT_FOUND,
            )

        if fallback_used:
            # include a debug hint so client logs can show why fallback was used
            company['_debug_fallback'] = True

        company = to_json_safe(company)
        if '_id' in company:
            company['id'] = company['_id']
        return Response(company, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def run_genetic_algorithm_variant2(request):
    """Run GA3 (Variant2) with company scoping"""
    try:
        config = request.data.get('config', {})
        session_id = request.data.get('session_id')
        if not session_id:
            return Response({'error': 'session_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        company_id = derive_company_id(request)
        if not company_id:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        dry_run = request.data.get('dry_run', False)
        # Ensure session belongs to current company
        session = InterviewSession.find_one({'_id': ObjectId(str(session_id)), 'company_id': company_id})
        if not session:
            return Response({'error': 'Session not found for this company'}, status=status.HTTP_404_NOT_FOUND)

        applicants = normalize_data(InterviewSession.get_session_applicants(session_id))
        interviewers = normalize_data(InterviewSession.get_session_interviewers(session_id))
        rooms = normalize_data(InterviewSession.get_session_rooms(session_id))
        if not applicants or not interviewers or not rooms:
            return Response({'error': 'Insufficient data.'}, status=status.HTTP_400_BAD_REQUEST)
        start = time.time()
        ga3 = GeneticAlgorithmVariant2(config=config)
        result = ga3.evolve(applicants, interviewers, rooms)
        chrom = result.get('best_solution')
        schedule_data = []
        if chrom and hasattr(chrom, 'genes'):
            for gene in chrom.genes:
                schedule_data.append({
                    'applicant_id': gene.applicant_id,
                    'interviewer_id': gene.interviewer_id,
                    'room_id': gene.room_id,
                    'start_time': gene.start_time.isoformat(),
                    'end_time': gene.end_time.isoformat(),
                    'position': gene.position,
                    'interview_date': gene.start_time.date().isoformat(),
                    'company_id': company_id
                })
        # session constraints only if we persist
        schedule_ids = []
        if schedule_data and not dry_run:
            for e in schedule_data:
                e['session_id'] = session_id
            Schedule.delete_all({'session_id': session_id})
            for entry in schedule_data:
                sid = Schedule.create(entry)
                schedule_ids.append(str(sid))
        result_doc = {
            'algorithm': 'GA3',
            'fitness_score': result.get('final_fitness', 0),
            'conflict_score': chrom.conflict_score if chrom else 0,
            'idle_time_score': chrom.idle_time_score if chrom else 0,
            'fairness_score': chrom.fairness_score if chrom else 0,
            'matching_score': chrom.matching_score if chrom else 0,
            'room_usage_score': chrom.room_usage_score if chrom else 0,
            'execution_time': time.time() - start,
            'generations': result.get('generations'),
            'fitness_history': result.get('fitness_history', []),
            'schedule_data': schedule_data,
            'schedule_ids': schedule_ids,
            'config_used': config,
            'dry_run': dry_run,
            'company_id': company_id,
            'session_id': session_id,
            'created_at': datetime.now()
        }
        rid = ScheduleResult.create(result_doc)

        # Log algorithm run
        log_action(
            request,
            'RUN_ALGORITHM',
            company_id=company_id,
            resource_type='session',
            resource_id=session_id,
            details={
                'algorithm': 'GA3',
                'dry_run': dry_run,
                'schedule_ids': schedule_ids,
            },
        )

        return Response(to_json_safe({'id': rid, **result_doc}))
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def run_genetic_algorithm_variant3(request):
    """Run GA4 (Variant3) with local search and company scoping"""
    try:
        config = request.data.get('config', {})
        session_id = request.data.get('session_id')
        if not session_id:
            return Response({'error': 'session_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        company_id = derive_company_id(request)
        if not company_id:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        dry_run = request.data.get('dry_run', False)
        # Ensure session belongs to current company
        session = InterviewSession.find_one({'_id': ObjectId(str(session_id)), 'company_id': company_id})
        if not session:
            return Response({'error': 'Session not found for this company'}, status=status.HTTP_404_NOT_FOUND)

        applicants = normalize_data(InterviewSession.get_session_applicants(session_id))
        interviewers = normalize_data(InterviewSession.get_session_interviewers(session_id))
        rooms = normalize_data(InterviewSession.get_session_rooms(session_id))
        if not applicants or not interviewers or not rooms:
            return Response({'error': 'Insufficient data.'}, status=status.HTTP_400_BAD_REQUEST)
        start = time.time()
        ga4 = GeneticAlgorithmVariant3(config=config)
        result = ga4.evolve(applicants, interviewers, rooms)
        chrom = result.get('best_solution')
        schedule_data = []
        if chrom and hasattr(chrom, 'genes'):
            for gene in chrom.genes:
                schedule_data.append({
                    'applicant_id': gene.applicant_id,
                    'interviewer_id': gene.interviewer_id,
                    'room_id': gene.room_id,
                    'start_time': gene.start_time.isoformat(),
                    'end_time': gene.end_time.isoformat(),
                    'position': gene.position,
                    'interview_date': gene.start_time.date().isoformat(),
                    'company_id': company_id
                })
        schedule_ids = []
        if schedule_data and not dry_run:
            for e in schedule_data:
                e['session_id'] = session_id
            Schedule.delete_all({'session_id': session_id})
            for entry in schedule_data:
                sid = Schedule.create(entry)
                schedule_ids.append(str(sid))
        result_doc = {
            'algorithm': 'GA4',
            'fitness_score': result.get('final_fitness', 0),
            'conflict_score': chrom.conflict_score if chrom else 0,
            'idle_time_score': chrom.idle_time_score if chrom else 0,
            'fairness_score': chrom.fairness_score if chrom else 0,
            'matching_score': chrom.matching_score if chrom else 0,
            'room_usage_score': chrom.room_usage_score if chrom else 0,
            'execution_time': time.time() - start,
            'generations': result.get('generations'),
            'fitness_history': result.get('fitness_history', []),
            'schedule_data': schedule_data,
            'schedule_ids': schedule_ids,
            'config_used': config,
            'dry_run': dry_run,
            'company_id': company_id,
            'session_id': session_id,
            'created_at': datetime.now()
        }
        rid = ScheduleResult.create(result_doc)

        # Log algorithm run (variant 3)
        log_action(
            request,
            'RUN_ALGORITHM',
            company_id=company_id,
            resource_type='session',
            resource_id=session_id,
            details={
                'algorithm': 'GA4',
                'dry_run': dry_run,
                'schedule_ids': schedule_ids,
            },
        )

        return Response(to_json_safe({'id': rid, **result_doc}))
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def import_excel(request):
    """Import data from Excel file"""
    try:
        file = request.FILES.get('file')
        data_type = request.data.get('type')  # 'applicants', 'interviewers', 'rooms', 'all'
        session_id = request.data.get('session_id')
        company_id = derive_company_id(request)

        if not company_id:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        created_ids = { 'applicants': [], 'interviewers': [], 'rooms': [] }
        total_counts = { 'applicants': 0, 'interviewers': 0, 'rooms': 0 }

        if data_type == 'all':
            # Expect sheets named Applicants, Interviewers, Rooms
            xls = pd.ExcelFile(file)
            sheets = {name.lower(): name for name in xls.sheet_names}

            if 'applicants' in sheets:
                df = pd.read_excel(xls, sheets['applicants'])
                records = df.to_dict('records')
                for r in records:
                    r['company_id'] = r.get('company_id') or company_id
                created_ids['applicants'] = Applicant.bulk_create(records)
                total_counts['applicants'] = len(records)
            if 'interviewers' in sheets:
                df = pd.read_excel(xls, sheets['interviewers'])
                records = df.to_dict('records')
                for r in records:
                    r['company_id'] = r.get('company_id') or company_id
                created_ids['interviewers'] = Interviewer.bulk_create(records)
                total_counts['interviewers'] = len(records)
            if 'rooms' in sheets:
                df = pd.read_excel(xls, sheets['rooms'])
                records = df.to_dict('records')
                for r in records:
                    r['company_id'] = r.get('company_id') or company_id
                created_ids['rooms'] = Room.bulk_create(records)
                total_counts['rooms'] = len(records)
        elif data_type in ('applicants','interviewers','rooms'):
            df = pd.read_excel(file)
            records = df.to_dict('records')
            for r in records:
                r['company_id'] = r.get('company_id') or company_id
            if data_type == 'applicants':
                created_ids['applicants'] = Applicant.bulk_create(records)
                total_counts['applicants'] = len(records)
            elif data_type == 'interviewers':
                created_ids['interviewers'] = Interviewer.bulk_create(records)
                total_counts['interviewers'] = len(records)
            elif data_type == 'rooms':
                created_ids['rooms'] = Room.bulk_create(records)
                total_counts['rooms'] = len(records)
        else:
            return Response({'error': 'Invalid data type'}, status=status.HTTP_400_BAD_REQUEST)

        # If session_id provided, associate created IDs with the session
        if session_id:
            coll = InterviewSession.get_collection()
            update = {}
            if created_ids['applicants']:
                update.setdefault('$addToSet', {}).setdefault('applicant_ids', {'$each': created_ids['applicants']})
            if created_ids['interviewers']:
                update.setdefault('$addToSet', {}).setdefault('interviewer_ids', {'$each': created_ids['interviewers']})
            if created_ids['rooms']:
                update.setdefault('$addToSet', {}).setdefault('room_ids', {'$each': created_ids['rooms']})
            if update:
                coll.update_one({'_id': ObjectId(session_id)}, update)

        response_data = {
            'message': 'Import completed',
            'counts': total_counts,
            'session_updated': bool(session_id)
        }

        # Log import action
        log_action(
            request,
            'IMPORT_EXCEL',
            company_id=company_id,
            resource_type='session' if session_id else None,
            resource_id=session_id,
            details={
                'data_type': data_type,
                'total_counts': total_counts,
            },
        )

        return Response(response_data)
    
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def export_schedules(request):
    """Export schedules to a formatted Excel workbook.
    Sheets:
      - First: "Danh sách ứng viên" for the session
      - Then: one sheet per (date, period) with horizontal column blocks per position
    Filename: "{session_name}__{start}_{end}__{company_name}.xlsx"
    """
    try:
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        from datetime import datetime as _dt

        # Helpers ------------------------------------------------------------
        def vn_weekday(dt: _dt) -> str:
            # Monday=0 ... Sunday=6
            names = [
                'Thứ hai', 'Thứ ba', 'Thứ tư', 'Thứ năm', 'Thứ sáu', 'Thứ bảy', 'Chủ nhật'
            ]
            return names[dt.weekday()]

        def period_of(dt: _dt) -> str:
            h = dt.hour
            if h < 12:
                return 'Sáng'
            if h < 18:
                return 'Chiều'
            return 'Tối'

        def fmt_span(st: _dt, et: _dt) -> str:
            return f"{st:%H}h{st:%M}m{st:%S}s - {et:%H}h{et:%M}m{et:%S}s"

        def safe_parse(iso_str: str) -> _dt:
            try:
                return _dt.fromisoformat(iso_str.replace('Z', '+00:00')).replace(tzinfo=None)
            except Exception:
                # Fallback: try without T
                try:
                    return _dt.strptime(iso_str, "%Y-%m-%d %H:%M:%S")
                except Exception:
                    return _dt.strptime(iso_str, "%Y-%m-%d")

        def sanitize_sheet_title(title: str) -> str:
            # Excel forbids: : \\/ ? * [ ] and limit 31 chars
            for ch in [":", "\\", "/", "?", "*", "[", "]"]:
                title = title.replace(ch, "-")
            # Collapse spaces
            title = " ".join(title.split())
            # Trim
            return title[:31]

        # Resolve session ----------------------------------------------------
        session_id = request.query_params.get('session_id')
        company_id = derive_company_id(request)

        session_doc = None
        if session_id:
            # Validate the session belongs to the caller's company
            session_doc = InterviewSession.find_one({'_id': ObjectId(str(session_id)), 'company_id': company_id})
            if not session_doc:
                return Response({'error': 'Session not found for this company'}, status=status.HTTP_404_NOT_FOUND)
        else:
            # Best effort: export for active session of this company
            active = _get_active_session_or_none(company_id)
            if active:
                session_doc = active
                session_id = str(active.get('_id'))

        if not session_doc:
            return Response({'error': 'Session not found or not specified'}, status=status.HTTP_400_BAD_REQUEST)

        # Data lookups scoped to session ------------------------------------
        schedules = Schedule.find_all({'session_id': session_id})
        if not schedules:
            return Response({'error': 'No schedules to export'}, status=status.HTTP_404_NOT_FOUND)

        applicants = InterviewSession.get_session_applicants(session_id)
        interviewers = InterviewSession.get_session_interviewers(session_id)
        rooms = InterviewSession.get_session_rooms(session_id)
        positions_map = Position.get_all_position_names()  # code -> name

        A = {str(a.get('_id')): a for a in applicants}
        I = {str(i.get('_id')): i for i in interviewers}
        R = {str(r.get('_id')): r for r in rooms}

        # Workbook -----------------------------------------------------------
        wb = Workbook()
        ws_list = wb.active
        ws_list.title = 'Danh sách ứng viên'

        # Header for applicant list
        ws_list.append(['Nhóm', 'Họ tên', 'MSSV'])
        for cell in ws_list[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

        # Rows for applicants
        for a in sorted(applicants, key=lambda x: (x.get('position',''), x.get('full_name',''))):
            ws_list.append([
                positions_map.get(a.get('position')) or a.get('position'),
                a.get('full_name'),
                a.get('student_id') or ''
            ])

        # Autosize columns
        for col in range(1, 4):
            ws_list.column_dimensions[get_column_letter(col)].width = [12, 28, 14][col-1]

        # Group schedules by (date, period) ---------------------------------
        groups = {}
        for s in schedules:
            st = safe_parse(s.get('start_time'))
            ed = safe_parse(s.get('end_time'))
            date_key = st.date().isoformat()
            group_key = (date_key, period_of(st))
            entry = {
                'st': st,
                'ed': ed,
                'room': R.get(str(s.get('room_id'))),
                'app': A.get(str(s.get('applicant_id'))),
                'intv': I.get(str(s.get('interviewer_id'))),
                'pos': s.get('position') or (A.get(str(s.get('applicant_id')) or '') or {}).get('position'),
                'note': s.get('note', ''),
            }
            groups.setdefault(group_key, []).append(entry)

        # For each group -> one sheet
        border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD'),
        )

        def write_block(ws, start_col, rows_for_pos, pos_code):
            headers = [
                'Số thứ tự',
                'Thời gian phỏng vấn',
                'Phòng',
                'Ứng viên',
                'Nhóm (Vị trí phỏng vấn)',
                'MSSV',
                'Người phỏng vấn',
                'Ghi chú',
            ]
            # Header row
            for idx, h in enumerate(headers):
                c = ws.cell(row=1, column=start_col + idx, value=h)
                c.font = Font(bold=True)
                c.fill = PatternFill(start_color="E5F2FF", end_color="E5F2FF", fill_type="solid")
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = border
            # Rows
            for r_idx, item in enumerate(rows_for_pos, start=2):
                ws.cell(row=r_idx, column=start_col + 0, value=r_idx - 1).border = border
                ws.cell(row=r_idx, column=start_col + 1, value=fmt_span(item['st'], item['ed'])).border = border
                ws.cell(row=r_idx, column=start_col + 2, value=(item['room'] or {}).get('room_code') or (item['room'] or {}).get('room_name') or '').border = border
                ws.cell(row=r_idx, column=start_col + 3, value=(item['app'] or {}).get('full_name') or '').border = border
                ws.cell(row=r_idx, column=start_col + 4, value=positions_map.get(pos_code) or pos_code).border = border
                ws.cell(row=r_idx, column=start_col + 5, value=(item['app'] or {}).get('student_id') or '').border = border
                ws.cell(row=r_idx, column=start_col + 6, value=(item['intv'] or {}).get('full_name') or '').border = border
                ws.cell(row=r_idx, column=start_col + 7, value=item.get('note','')).border = border

            # Widths
            widths = [10, 24, 12, 26, 20, 12, 22, 26]
            for i, w in enumerate(widths):
                ws.column_dimensions[get_column_letter(start_col + i)].width = w

        used_titles = set()
        for (date_key, per), entries in sorted(groups.items(), key=lambda x: (x[0][0], ['Sáng','Chiều','Tối'].index(x[0][1]) if x[0][1] in ['Sáng','Chiều','Tối'] else 3)):
            # Sheet name
            dt = _dt.strptime(date_key, "%Y-%m-%d")
            # Replace slashes to satisfy Excel constraints
            sheet_name = f"{per} - {vn_weekday(dt)} {dt:%d-%m-%Y}"
            sheet_name = sanitize_sheet_title(sheet_name)
            # Ensure uniqueness after sanitization/truncation
            base_name = sheet_name
            idx = 2
            while sheet_name in used_titles:
                suffix = f" ({idx})"
                sheet_name = (base_name[: (31 - len(suffix))] + suffix) if len(base_name) + len(suffix) > 31 else base_name + suffix
                idx += 1
            used_titles.add(sheet_name)
            ws = wb.create_sheet(title=sheet_name)

            # Partition by position code
            entries.sort(key=lambda e: (e['st'], (positions_map.get(e['pos']) or e['pos'] or '')))
            pos_codes = []
            for e in entries:
                code = e['pos'] or 'Unknown'
                if code not in pos_codes:
                    pos_codes.append(code)

            start_col = 1
            gap = 2  # two blank columns between blocks
            for pcode in pos_codes:
                rows_for_pos = [x for x in entries if (x['pos'] or 'Unknown') == pcode]
                write_block(ws, start_col, rows_for_pos, pcode)
                start_col += 8 + gap

            ws.freeze_panes = 'A2'

        # Filename -----------------------------------------------------------
        company_name = ''
        if session_doc.get('company_id'):
            comp = Company.find_by_id(session_doc.get('company_id'))
            company_name = comp.get('name') if comp else ''
        period = f"{session_doc.get('start_date')}_{session_doc.get('end_date')}"
        # Use double underscore as segment separators per request
        raw_name = f"{session_doc.get('name') or session_doc.get('code') or 'Session'}__{period}__{company_name or 'Company'}".strip()

        # Sanitize filename for ASCII (avoid header encoding corruption)
        import unicodedata, re, urllib.parse
        def ascii_slug(s: str) -> str:
            # Normalize & strip diacritics
            nfkd = unicodedata.normalize('NFKD', s)
            without_diacritics = ''.join(c for c in nfkd if not unicodedata.combining(c))
            # Replace forbidden chars
            without_diacritics = re.sub(r'[\\/:*?"<>|]+', '-', without_diacritics)
            # Collapse whitespace
            without_diacritics = re.sub(r'\s+', ' ', without_diacritics).strip()
            # Replace spaces with underscores for safer cross-platform handling
            safe = without_diacritics.replace(' ', '_')
            # Keep only allowed chars
            safe = re.sub(r'[^A-Za-z0-9._-]+', '', safe)
            # Guard length
            return safe[:120] or 'export'
        ascii_name = ascii_slug(raw_name)
        utf8_quoted = urllib.parse.quote(raw_name + '.xlsx')

        # Response -----------------------------------------------------------
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # Provide both ASCII fallback and RFC 5987 UTF-8 filename*
        response['Content-Disposition'] = (
            f'attachment; filename="{ascii_name}.xlsx"; filename*=UTF-8''{utf8_quoted}'
        )

        # Log export action (best-effort, ignore failures)
        try:
            log_action(
                request,
                'EXPORT_SCHEDULE',
                company_id=company_id,
                resource_type='session',
                resource_id=str(session_doc.get('_id')) if session_doc else None,
                details={
                    'session_name': session_doc.get('name') if session_doc else None,
                    'start_date': session_doc.get('start_date') if session_doc else None,
                    'end_date': session_doc.get('end_date') if session_doc else None,
                    'file_name': raw_name,
                },
            )
        except Exception:
            pass

        return response

    except Exception as e:
        print(f"❌ Export error: {e}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def dashboard_stats(request):
    """Get dashboard statistics scoped by company and optionally session.

    Query params:
      - company_id: override company (defaults to derive_company_id)
      - session_id: when provided, counts are limited to that session
    """
    try:
        company_id = request.query_params.get('company_id') or derive_company_id(request)
        session_id = request.query_params.get('session_id')

        base_filter = {}
        if company_id:
            base_filter['company_id'] = company_id
        if session_id:
            base_filter['session_id'] = session_id

        # Dynamic positions-based stats
        pos_filter = {'is_active': True}
        if company_id:
            pos_filter['company_id'] = company_id
        positions = Position.find_all(pos_filter)

        pos_counts_app = {}
        pos_counts_int = {}
        for p in positions:
            code = p.get('code')
            name = p.get('name')
            name_or_code = name or code
            app_filter = {'position': code}
            int_filter = {'position': code}
            if company_id:
                app_filter['company_id'] = company_id
                int_filter['company_id'] = company_id
            pos_counts_app[name_or_code] = Applicant.count(app_filter)
            pos_counts_int[name_or_code] = Interviewer.count(int_filter)

        app_filter_total = {}
        int_filter_total = {}
        room_filter_total = {}
        sched_filter_total = {}
        if company_id:
            app_filter_total['company_id'] = company_id
            int_filter_total['company_id'] = company_id
            room_filter_total['company_id'] = company_id
            sched_filter_total['company_id'] = company_id
        if session_id:
            sched_filter_total['session_id'] = session_id

        # ScheduleResult-based stats ("schedule options" / candidate schedules)
        sr_filter_total = {}
        if company_id:
            sr_filter_total['company_id'] = company_id
        if session_id:
            sr_filter_total['session_id'] = session_id

        stats = {
            'applicants': {
                'total': Applicant.count(app_filter_total or None),
                'by_position': pos_counts_app,
            },
            'interviewers': {
                'total': Interviewer.count(int_filter_total or None),
                'by_position': pos_counts_int,
            },
            'rooms': {
                'total': Room.count(room_filter_total or None),
            },
            'schedules': {
                # Number of generated schedule options (ScheduleResult documents)
                'total': ScheduleResult.count(sr_filter_total or None),
            },
        }
        return Response(stats)
    except Exception as e:
        import traceback
        print(f"Error in dashboard_stats: {str(e)}")
        print(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CompanyAPIView(APIView):
    """API endpoint for Companies (multi-tenant)"""

    def get(self, request, pk=None):
        # Only allow access to the authenticated user's company
        cid = derive_company_id(request)
        if not cid:
            return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)
        if pk:
            if str(pk) != str(cid):
                return Response({'error': 'Company not found'}, status=status.HTTP_404_NOT_FOUND)
            # Bypass tenant scoping when loading the company document itself
            company = Company.find_by_id(pk, company_id=None)
            if company:
                return Response(company)
            return Response({'error': 'Company not found'}, status=status.HTTP_404_NOT_FOUND)
        # Return the current company document (do not filter by company_id)
        company = Company.find_by_id(cid, company_id=None)
        return Response([company] if company else [])

    def post(self, request):
        # Regular users cannot create new companies via this endpoint
        return Response({'error': 'Not allowed'}, status=status.HTTP_403_FORBIDDEN)

    def put(self, request, pk):
        cid = derive_company_id(request)
        if not cid:
            return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)
        if str(pk) != str(cid):
            return Response({'error': 'Company not found'}, status=status.HTTP_404_NOT_FOUND)
        if Company.update(pk, request.data):
            updated = Company.find_by_id(pk)
            # Log company update
            log_action(
                request,
                'UPDATE_COMPANY',
                company_id=str(pk),
                resource_type='company',
                resource_id=str(pk),
                details={
                    'changes': request.data,
                },
            )
            return Response(updated)
        return Response({'error': 'Company not found'}, status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, pk):
        # Disallow destructive action for safety
        return Response({'error': 'Not allowed'}, status=status.HTTP_403_FORBIDDEN)


@api_view(['POST'])
def register_user(request):
    """Simple registration: username, password, company_code or company_id.
    If company_code provided and company doesn't exist -> create it.
    Returns auth token.
    """
    try:
        username = request.data.get('username')
        password = request.data.get('password')
        company_code = request.data.get('company_code')
        company_id = request.data.get('company_id')
        if not username or not password:
            return Response({'error': 'username and password required'}, status=status.HTTP_400_BAD_REQUEST)
        # enforce unique username
        if User.find_one({'username': username}):
            return Response({'error': 'Username already exists'}, status=status.HTTP_400_BAD_REQUEST)
        # resolve company
        company_doc = None
        if company_id:
            company_doc = Company.find_by_id(company_id)
            if not company_doc:
                return Response({'error': 'Invalid company_id'}, status=status.HTTP_400_BAD_REQUEST)
        elif company_code:
            company_doc = Company.find_one({'code': company_code})
            if not company_doc:
                # auto-create company with name=code
                new_company_id = Company.create({'name': company_code, 'code': company_code, 'created_at': datetime.now()})
                company_doc = Company.find_by_id(new_company_id)
        else:
            return Response({'error': 'company_code or company_id required'}, status=status.HTTP_400_BAD_REQUEST)
        hashed = User.hash_password(password)
        token = User.generate_token()
        # New users default to manager role; system/admin users are seeded manually
        user_doc = {
            'username': username,
            'password': hashed,
            'company_id': company_doc.get('_id'),
            'token': token,
            'role': request.data.get('role') or 'manager',
            'created_at': datetime.now()
        }
        user_id = User.create(user_doc)

        # Log user creation
        log_action(
            request,
            'CREATE_USER',
            company_id=str(company_doc.get('_id')),
            resource_type='user',
            resource_id=str(user_id),
            details={
                'username': username,
                'role': user_doc.get('role'),
            },
        )

        return Response({
            'id': str(user_id),
            'username': username,
            'company_id': str(company_doc.get('_id')),
            'token': token,
            'role': user_doc.get('role'),
        }, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def login_user(request):
    """Login: username + password -> returns token (rotated)."""
    try:
        username = request.data.get('username')
        password = request.data.get('password')
        if not username or not password:
            return Response({'error': 'username and password required'}, status=status.HTTP_400_BAD_REQUEST)
        user = User.find_one({'username': username})
        if not user or not User.verify_password(user.get('password', ''), password):
            return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)
        # rotate token
        new_token = User.generate_token()
        User.update(user.get('_id'), {'token': new_token})
        cid = user.get('company_id')
        return Response({
            'id': str(user.get('_id')),
            'username': username,
            'company_id': str(cid) if cid is not None else None,
            'token': new_token,
            'role': user.get('role'),
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def run_genetic_algorithm(request):
    """Run Genetic Algorithm for scheduling"""
    try:
        config = request.data.get('config', {})
        session_id = request.data.get('session_id')
        if not session_id:
            return Response({'error': 'session_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        company_id = derive_company_id(request)
        if not company_id:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Ensure session belongs to current company
        session = InterviewSession.find_one({'_id': ObjectId(str(session_id)), 'company_id': company_id})
        if not session:
            return Response({'error': 'Session not found for this company'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get data from MongoDB
        # Select only entities belonging to the chosen session
        applicants = normalize_data(InterviewSession.get_session_applicants(session_id))
        interviewers = normalize_data(InterviewSession.get_session_interviewers(session_id))
        rooms = normalize_data(InterviewSession.get_session_rooms(session_id))
        
        if not applicants or not interviewers or not rooms:
            return Response({
                'error': 'Insufficient data. Need applicants, interviewers, and rooms.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Run GA
        start_time = time.time()
        ga = GeneticAlgorithm(config=config)
        result = ga.evolve(applicants, interviewers, rooms)
        
        # Extract results
        best_chromosome = result.get('best_solution')
        best_fitness = result.get('final_fitness', 0)
        generations = result.get('generations', 0)
        execution_time = time.time() - start_time
        
        print(f"📊 Views: best_chromosome = {best_chromosome}")
        print(f"📊 Views: has genes attr = {hasattr(best_chromosome, 'genes') if best_chromosome else False}")
        if best_chromosome and hasattr(best_chromosome, 'genes'):
            print(f"📊 Views: genes count = {len(best_chromosome.genes)}")
        
        # Convert Chromosome genes to schedule data
        schedule_data = []
        if best_chromosome and hasattr(best_chromosome, 'genes'):
            print(f"📊 Views: Starting conversion loop...")
            for i, gene in enumerate(best_chromosome.genes):
                try:
                    entry = {
                        'applicant_id': gene.applicant_id,
                        'interviewer_id': gene.interviewer_id,
                        'room_id': gene.room_id,
                        'start_time': gene.start_time.isoformat() if hasattr(gene.start_time, 'isoformat') else str(gene.start_time),
                        'end_time': gene.end_time.isoformat() if hasattr(gene.end_time, 'isoformat') else str(gene.end_time),
                        'position': gene.position,
                        'interview_date': gene.start_time.date().isoformat() if hasattr(gene.start_time, 'date') else None
                    }
                    schedule_data.append(entry)
                    if i == 0:
                        print(f"📊 Views: First gene converted: {entry}")
                except Exception as e:
                    print(f"📊 Views: Error converting gene {i}: {e}")
            print(f"📊 Views: Conversion complete. schedule_data length = {len(schedule_data)}")
        
        # Apply session-aware constraints and save schedules
        schedule_ids = []
        if schedule_data:
            active_session = _get_active_session_or_none(company_id)
            if active_session:
                for e in schedule_data:
                    e['session_id'] = active_session.get('_id')
            filtered_data, stats = _filter_by_session_constraints(schedule_data, active_session)
            print(f"🧹 [GA] Filtered schedules: kept={len(filtered_data)}, skipped_conflicts={stats['skipped_conflicts']}, skipped_overtime={stats['skipped_overtime']}")
            schedule_data = filtered_data

        if schedule_data:
            print(f"💾 Saving {len(schedule_data)} schedules to database...")
            # Clear old schedules in current session only (if any)
            Schedule.delete_all({'session_id': session_id})
            
            for schedule_entry in schedule_data:
                schedule_entry['session_id'] = session_id
                if company_id:
                    schedule_entry['company_id'] = company_id
                schedule_id = Schedule.create(schedule_entry)
                schedule_ids.append(str(schedule_id))
            print(f"✅ Saved {len(schedule_ids)} schedules")
        
        # Save result
        result_data = {
            'algorithm': 'GA',
            'fitness_score': best_fitness,
            'conflict_score': best_chromosome.conflict_score if best_chromosome else 0,
            'idle_time_score': best_chromosome.idle_time_score if best_chromosome else 0,
            'fairness_score': best_chromosome.fairness_score if best_chromosome else 0,
            'matching_score': best_chromosome.matching_score if best_chromosome else 0,
            'room_usage_score': best_chromosome.room_usage_score if best_chromosome else 0,
            'execution_time': execution_time,
            'generations': generations,
            'schedule_data': schedule_data,
            'schedule_ids': schedule_ids,
            'config_used': config,
            'company_id': company_id,
            'session_id': session_id,
            'created_at': datetime.now()
        }
        
        result_id = ScheduleResult.create(result_data)
        # Return JSON-safe payload without nested ObjectId by avoiding re-fetch
        safe_result = {
            'id': result_id,
            **result_data
        }
        
        return Response(to_json_safe(safe_result))
    
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def run_genetic_algorithm_variant(request):
    """Run Genetic Algorithm Variant (GA2) for scheduling"""
    try:
        config = request.data.get('config', {})
        session_id = request.data.get('session_id')
        if not session_id:
            return Response({'error': 'session_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        company_id = derive_company_id(request)
        if not company_id:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Ensure session belongs to current company
        session = InterviewSession.find_one({'_id': ObjectId(str(session_id)), 'company_id': company_id})
        if not session:
            return Response({'error': 'Session not found for this company'}, status=status.HTTP_404_NOT_FOUND)
        
        applicants = normalize_data(InterviewSession.get_session_applicants(session_id))
        interviewers = normalize_data(InterviewSession.get_session_interviewers(session_id))
        rooms = normalize_data(InterviewSession.get_session_rooms(session_id))

        if not applicants or not interviewers or not rooms:
            return Response({'error': 'Insufficient data. Need applicants, interviewers, and rooms.'}, status=status.HTTP_400_BAD_REQUEST)

        start_time = time.time()
        ga2 = GeneticAlgorithmVariant(config=config)
        result = ga2.evolve(applicants, interviewers, rooms)

        best_chromosome = result.get('best_solution')
        best_fitness = result.get('final_fitness', 0)
        generations = result.get('generations', 0)
        execution_time = time.time() - start_time

        schedule_data = []
        if best_chromosome and hasattr(best_chromosome, 'genes'):
            for gene in best_chromosome.genes:
                schedule_data.append({
                    'applicant_id': gene.applicant_id,
                    'interviewer_id': gene.interviewer_id,
                    'room_id': gene.room_id,
                    'start_time': gene.start_time.isoformat() if hasattr(gene.start_time, 'isoformat') else str(gene.start_time),
                    'end_time': gene.end_time.isoformat() if hasattr(gene.end_time, 'isoformat') else str(gene.end_time),
                    'position': gene.position,
                    'interview_date': gene.start_time.date().isoformat() if hasattr(gene.start_time, 'date') else None
                })

        # Apply session-aware constraints and save
        schedule_ids = []
        if schedule_data:
            active_session = _get_active_session_or_none(company_id)
            if active_session:
                for e in schedule_data:
                    e['session_id'] = active_session.get('_id')
            filtered_data, stats = _filter_by_session_constraints(schedule_data, active_session)
            print(f"🧹 [GA2] Filtered schedules: kept={len(filtered_data)}, skipped_conflicts={stats['skipped_conflicts']}, skipped_overtime={stats['skipped_overtime']}")
            schedule_data = filtered_data

        if schedule_data:
            print(f"💾 [GA2] Saving {len(schedule_data)} schedules to database...")
            Schedule.delete_all({'session_id': session_id})
            for schedule_entry in schedule_data:
                schedule_entry['session_id'] = session_id
                if company_id:
                    schedule_entry['company_id'] = company_id
                schedule_id = Schedule.create(schedule_entry)
                schedule_ids.append(str(schedule_id))
            print(f"✅ [GA2] Saved {len(schedule_ids)} schedules")

        result_data = {
            'algorithm': 'GA2',
            'fitness_score': best_fitness,
            'conflict_score': best_chromosome.conflict_score if best_chromosome else 0,
            'idle_time_score': best_chromosome.idle_time_score if best_chromosome else 0,
            'fairness_score': best_chromosome.fairness_score if best_chromosome else 0,
            'matching_score': best_chromosome.matching_score if best_chromosome else 0,
            'room_usage_score': best_chromosome.room_usage_score if best_chromosome else 0,
            'execution_time': execution_time,
            'generations': generations,
            'schedule_data': schedule_data,
            'schedule_ids': schedule_ids,
            'config_used': config,
            'company_id': company_id,
            'session_id': session_id,
            'created_at': datetime.now()
        }

        result_id = ScheduleResult.create(result_data)
        safe_result = {'id': result_id, **result_data}
        return Response(to_json_safe(safe_result))
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




@api_view(['GET'])
def algorithm_results(request):
    """Get algorithm results.

    Optional query params:
      - session_id: filter by session
      - top: integer; if provided, return top-N by fitness_score (descending)
      - selected: 'true'|'false' to filter by is_selected flag
    """
    try:
        company_id = derive_company_id(request)
        session_id = request.query_params.get('session_id')
        top = request.query_params.get('top')
        selected = request.query_params.get('selected')

        filt = {}
        if company_id:
            filt['company_id'] = company_id
        if session_id:
            filt['session_id'] = session_id
        if selected is not None:
            if selected.lower() in ('true', '1', 'yes'):
                filt['is_selected'] = True
            elif selected.lower() in ('false', '0', 'no'):
                filt['is_selected'] = False

        limit = None
        try:
            if top is not None:
                limit = int(top)
        except ValueError:
            limit = None

        # Default limit for non-top queries
        if limit is None:
            limit = 20

        results = ScheduleResult.find_all(
            filter_dict=filt or None,
            limit=limit,
            sort=[("fitness_score", -1), ("created_at", -1)],
        )
        return Response(to_json_safe(results))
    except Exception as e:
        print(f"❌ Error in algorithm_results: {e}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def compare_algorithms(request):
    """Compare GA family variants (GA, GA2, GA3, GA4)"""
    try:
        config = request.data.get('config', {})
        session_id = request.data.get('session_id')
        if not session_id:
            return Response({'error': 'session_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        company_id = derive_company_id(request)
        if not company_id:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Ensure session belongs to current company
        session = InterviewSession.find_one({'_id': ObjectId(str(session_id)), 'company_id': company_id})
        if not session:
            return Response({'error': 'Session not found for this company'}, status=status.HTTP_404_NOT_FOUND)

        # Session-scoped data only
        applicants = normalize_data(InterviewSession.get_session_applicants(session_id))
        interviewers = normalize_data(InterviewSession.get_session_interviewers(session_id))
        rooms = normalize_data(InterviewSession.get_session_rooms(session_id))
        
        if not applicants or not interviewers or not rooms:
            return Response({
                'error': 'Insufficient data. Need applicants, interviewers, and rooms.'
            }, status=status.HTTP_400_BAD_REQUEST)
        comparison_results = []
        
        # 1. Run Genetic Algorithm
        print("🔄 Running Genetic Algorithm...")
        ga_start = time.time()
        ga_config = config.get('GA', {
            'POPULATION_SIZE': 100,
            'GENERATIONS': 200,
            'CROSSOVER_RATE': 0.8,
            'MUTATION_RATE': 0.15,
            'WEIGHTS': {
                'conflict': 0.4,
                'idle_time': 0.2,
                'fairness': 0.2,
                'matching': 0.1,
                'room_usage': 0.1
            }
        })
        ga = GeneticAlgorithm(config=ga_config)
        ga_result = ga.evolve(applicants, interviewers, rooms)
        ga_time = time.time() - ga_start
        
        ga_chromosome = ga_result.get('best_solution')
        # resource metrics
        ga_interviewer_load = {}
        ga_room_util = {}
        if ga_chromosome and hasattr(ga_chromosome, 'genes'):
            for g in ga_chromosome.genes:
                ga_interviewer_load[g.interviewer_id] = ga_interviewer_load.get(g.interviewer_id, 0) + 1
                ga_room_util[g.room_id] = ga_room_util.get(g.room_id, 0) + 1
        comparison_results.append({
            'algorithm': 'GA',
            'fitness_score': ga_result.get('final_fitness', 0),
            'execution_time': ga_time,
            'schedules_count': len(ga_chromosome.genes) if ga_chromosome else 0,
            'conflict_score': ga_chromosome.conflict_score if ga_chromosome else 0,
            'idle_time_score': ga_chromosome.idle_time_score if ga_chromosome else 0,
            'fairness_score': ga_chromosome.fairness_score if ga_chromosome else 0,
            'matching_score': ga_chromosome.matching_score if ga_chromosome else 0,
            'room_usage_score': ga_chromosome.room_usage_score if ga_chromosome else 0,
            'interviewer_load': ga_interviewer_load,
            'room_utilization': ga_room_util,
        })

        # 1b. Run Genetic Algorithm Variant (GA2)
        print("🔄 Running Genetic Algorithm Variant (GA2)...")
        ga2_start = time.time()
        ga2_config = config.get('GA2', {
            'POPULATION_SIZE': 100,
            'GENERATIONS': 200,
            'CROSSOVER_RATE': 0.9,
            'MUTATION_RATE': 0.2,
            'ELITISM_RATE': 0.05,
        })
        ga2 = GeneticAlgorithmVariant(config=ga2_config)
        ga2_result = ga2.evolve(applicants, interviewers, rooms)
        ga2_time = time.time() - ga2_start

        ga2_chromosome = ga2_result.get('best_solution')
        ga2_interviewer_load = {}
        ga2_room_util = {}
        if ga2_chromosome and hasattr(ga2_chromosome, 'genes'):
            for g in ga2_chromosome.genes:
                ga2_interviewer_load[g.interviewer_id] = ga2_interviewer_load.get(g.interviewer_id, 0) + 1
                ga2_room_util[g.room_id] = ga2_room_util.get(g.room_id, 0) + 1
        comparison_results.append({
            'algorithm': 'GA2',
            'fitness_score': ga2_result.get('final_fitness', 0),
            'execution_time': ga2_time,
            'schedules_count': len(ga2_chromosome.genes) if ga2_chromosome else 0,
            'conflict_score': ga2_chromosome.conflict_score if ga2_chromosome else 0,
            'idle_time_score': ga2_chromosome.idle_time_score if ga2_chromosome else 0,
            'fairness_score': ga2_chromosome.fairness_score if ga2_chromosome else 0,
            'matching_score': ga2_chromosome.matching_score if ga2_chromosome else 0,
            'room_usage_score': ga2_chromosome.room_usage_score if ga2_chromosome else 0,
            'interviewer_load': ga2_interviewer_load,
            'room_utilization': ga2_room_util,
        })
        
        # 2. Run GA3
        print("🔄 Running Genetic Algorithm Variant 2 (GA3)...")
        ga3_start = time.time()
        ga3_config = config.get('GA3', {
            'POPULATION_SIZE': 120,
            'GENERATIONS': 250,
            'CROSSOVER_RATE': 0.85,
            'MUTATION_RATE': 0.12
        })
        ga3 = GeneticAlgorithmVariant2(config=ga3_config)
        ga3_result = ga3.evolve(applicants, interviewers, rooms)
        ga3_time = time.time() - ga3_start
        ga3_chromosome = ga3_result.get('best_solution')
        ga3_interviewer_load = {}
        ga3_room_util = {}
        if ga3_chromosome and hasattr(ga3_chromosome, 'genes'):
            for g in ga3_chromosome.genes:
                ga3_interviewer_load[g.interviewer_id] = ga3_interviewer_load.get(g.interviewer_id, 0) + 1
                ga3_room_util[g.room_id] = ga3_room_util.get(g.room_id, 0) + 1
        comparison_results.append({
            'algorithm': 'GA3',
            'fitness_score': ga3_result.get('final_fitness', 0),
            'execution_time': ga3_time,
            'schedules_count': len(ga3_chromosome.genes) if ga3_chromosome else 0,
            'conflict_score': ga3_chromosome.conflict_score if ga3_chromosome else 0,
            'idle_time_score': ga3_chromosome.idle_time_score if ga3_chromosome else 0,
            'fairness_score': ga3_chromosome.fairness_score if ga3_chromosome else 0,
            'matching_score': ga3_chromosome.matching_score if ga3_chromosome else 0,
            'room_usage_score': ga3_chromosome.room_usage_score if ga3_chromosome else 0,
            'interviewer_load': ga3_interviewer_load,
            'room_utilization': ga3_room_util,
        })

        # 3. Run GA4
        print("🔄 Running Genetic Algorithm Variant 3 (GA4)...")
        ga4_start = time.time()
        ga4_config = config.get('GA4', {
            'POPULATION_SIZE': 120,
            'GENERATIONS': 250,
            'CROSSOVER_RATE': 0.9,
            'MUTATION_RATE': 0.18,
            'LOCAL_SEARCH_RATE': 0.3
        })
        ga4 = GeneticAlgorithmVariant3(config=ga4_config)
        ga4_result = ga4.evolve(applicants, interviewers, rooms)
        ga4_time = time.time() - ga4_start
        ga4_chromosome = ga4_result.get('best_solution')
        ga4_interviewer_load = {}
        ga4_room_util = {}
        if ga4_chromosome and hasattr(ga4_chromosome, 'genes'):
            for g in ga4_chromosome.genes:
                ga4_interviewer_load[g.interviewer_id] = ga4_interviewer_load.get(g.interviewer_id, 0) + 1
                ga4_room_util[g.room_id] = ga4_room_util.get(g.room_id, 0) + 1
        comparison_results.append({
            'algorithm': 'GA4',
            'fitness_score': ga4_result.get('final_fitness', 0),
            'execution_time': ga4_time,
            'schedules_count': len(ga4_chromosome.genes) if ga4_chromosome else 0,
            'conflict_score': ga4_chromosome.conflict_score if ga4_chromosome else 0,
            'idle_time_score': ga4_chromosome.idle_time_score if ga4_chromosome else 0,
            'fairness_score': ga4_chromosome.fairness_score if ga4_chromosome else 0,
            'matching_score': ga4_chromosome.matching_score if ga4_chromosome else 0,
            'room_usage_score': ga4_chromosome.room_usage_score if ga4_chromosome else 0,
            'interviewer_load': ga4_interviewer_load,
            'room_utilization': ga4_room_util,
        })
        
        print(f"✅ Comparison complete!")
        return Response({
            'results': comparison_results,
            'timestamp': datetime.now().isoformat(),
            'company_id': company_id,
            'session_id': session_id,
        })
        
    except Exception as e:
        print(f"❌ Error in compare_algorithms: {e}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def action_logs(request):
    """List action logs filtered by action type, scoped strictly to caller's company.

    Query params:
      - action_type: filter by action_type
      - limit: max number of records (default 100)
    """
    try:
        user = get_request_user(request)
        if not user:
            return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)

        role = user.get('role')
        if role not in ('admin', 'manager'):
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

        # Always scope to token-derived company (no client override)
        company_id = derive_company_id(request)
        if not company_id:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)

        action_type = request.query_params.get('action_type')
        try:
            limit = int(request.query_params.get('limit', 100))
        except ValueError:
            limit = 100

        query = {'company_id': company_id}
        if action_type:
            query['action_type'] = action_type

        logs = ActionLog.find_all(query, limit=limit, sort=[('created_at', -1)])
        for l in logs:
            if '_id' in l:
                l['id'] = str(l['_id'])
        return Response(to_json_safe(logs))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def generate_top_schedules(request):
    """Generate multiple GA family runs and return top-K without persisting schedules."""
    try:
        k = int(request.data.get('k', 5))
        runs = int(request.data.get('runs', 8))
        config = request.data.get('config', {})
        session_id = request.data.get('session_id')
        if not session_id:
            return Response({'error': 'session_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        company_id = derive_company_id(request)
        if not company_id:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        # Ensure session belongs to caller's company
        session = InterviewSession.find_one({'_id': ObjectId(str(session_id)), 'company_id': company_id})
        if not session:
            return Response({'error': 'Session not found for this company'}, status=status.HTTP_404_NOT_FOUND)
        applicants = normalize_data(InterviewSession.get_session_applicants(session_id))
        interviewers = normalize_data(InterviewSession.get_session_interviewers(session_id))
        rooms = normalize_data(InterviewSession.get_session_rooms(session_id))
        if not applicants or not interviewers or not rooms:
            return Response({'error': 'Insufficient data.'}, status=status.HTTP_400_BAD_REQUEST)
        variants = [
            ('GA', GeneticAlgorithm, config.get('GA', {})),
            ('GA2', GeneticAlgorithmVariant, config.get('GA2', {})),
            ('GA3', GeneticAlgorithmVariant2, config.get('GA3', {})),
            ('GA4', GeneticAlgorithmVariant3, config.get('GA4', {})),
        ]
        candidate_results = []
        for i in range(runs):
            name, cls, cfg = random.choice(variants)
            alg = cls(cfg)
            res = alg.evolve(applicants, interviewers, rooms)
            chrom = res.get('best_solution')
            schedule_data = []
            if chrom and hasattr(chrom, 'genes'):
                for gene in chrom.genes:
                    schedule_data.append({
                        'applicant_id': gene.applicant_id,
                        'interviewer_id': gene.interviewer_id,
                        'room_id': gene.room_id,
                        'start_time': gene.start_time.isoformat(),
                        'end_time': gene.end_time.isoformat(),
                        'position': gene.position,
                        'interview_date': gene.start_time.date().isoformat()
                    })
            doc = {
                'algorithm': name,
                'fitness_score': res.get('final_fitness', 0),
                'conflict_score': chrom.conflict_score if chrom else 0,
                'idle_time_score': chrom.idle_time_score if chrom else 0,
                'fairness_score': chrom.fairness_score if chrom else 0,
                'matching_score': chrom.matching_score if chrom else 0,
                'room_usage_score': chrom.room_usage_score if chrom else 0,
                'generations': res.get('generations'),
                'schedule_data': schedule_data,
                'fitness_history': res.get('fitness_history', []),
                'created_at': datetime.now(),
                'company_id': company_id,
                'session_id': session_id,
                'candidate_set': True,
                'is_selected': False
            }
            rid = ScheduleResult.create(doc)
            candidate_results.append({'id': rid, **doc})
        # pick top k by fitness
        candidate_results.sort(key=lambda x: x['fitness_score'], reverse=True)
        top_k = candidate_results[:k]
        return Response(to_json_safe({'top_k': top_k, 'total_generated': len(candidate_results), 'session_id': session_id, 'company_id': company_id}))
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def choose_schedule_result(request):
    """Select one candidate schedule result and persist its schedule_data as official schedules."""
    try:
        result_id = request.data.get('result_id')
        session_id = request.data.get('session_id')
        company_id = derive_company_id(request)
        if not result_id:
            return Response({'error': 'result_id required'}, status=status.HTTP_400_BAD_REQUEST)
        if not session_id:
            return Response({'error': 'session_id required'}, status=status.HTTP_400_BAD_REQUEST)
        if not company_id:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        # Validate session ownership
        session = InterviewSession.find_one({'_id': ObjectId(str(session_id)), 'company_id': company_id})
        if not session:
            return Response({'error': 'Session not found for this company'}, status=status.HTTP_404_NOT_FOUND)

        result = ScheduleResult.find_by_id(result_id)
        if not result:
            return Response({'error': 'Result not found'}, status=status.HTTP_404_NOT_FOUND)
        if str(result.get('company_id')) != str(company_id) or str(result.get('session_id')) != str(session_id):
            return Response({'error': 'Result does not belong to this company/session'}, status=status.HTTP_403_FORBIDDEN)
        schedule_data = result.get('schedule_data', [])
        # Clear existing schedules for provided session
        Schedule.delete_all({'session_id': session_id})
        saved_ids = []
        for entry in schedule_data:
            entry['session_id'] = session_id
            if company_id:
                entry['company_id'] = company_id
            # ensure interview_date present
            if 'interview_date' not in entry and 'start_time' in entry:
                try:
                    entry['interview_date'] = datetime.fromisoformat(entry['start_time']).date().isoformat()
                except Exception:
                    entry['interview_date'] = datetime.now().date().isoformat()
            sid = Schedule.create(entry)
            saved_ids.append(str(sid))
        # Mark this result as selected and others of same session/company as not selected
        coll = ScheduleResult.get_collection()
        coll.update_many(
            {'session_id': session_id, 'company_id': company_id},
            {'$set': {'is_selected': False}},
        )
        ScheduleResult.update(result_id, {'is_selected': True})
        return Response({
            'selected_result_id': result_id,
            'persisted_count': len(saved_ids),
            'company_id': company_id,
            'session_id': session_id,
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AlgorithmConfigAPIView(APIView):
    """API endpoint for Algorithm Configuration"""
    
    def get(self, request, pk=None):
        company_id = derive_company_id(request)
        if pk:
            config = AlgorithmConfig.find_by_id(pk)
            if config and (not company_id or str(config.get('company_id')) == str(company_id)):
                return Response(config)
            return Response({'error': 'Config not found'}, status=status.HTTP_404_NOT_FOUND)
        # List only configs belonging to caller's company
        filt = {'company_id': company_id} if company_id else {}
        configs = AlgorithmConfig.find_all(filt or None)
        return Response(configs)
    
    def post(self, request):
        data = request.data.copy()
        cid = derive_company_id(request)
        if not cid:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        # Force company scoping
        data['company_id'] = cid
        if not AlgorithmConfig.validate(data):
            return Response({'error': 'Invalid data'}, status=status.HTTP_400_BAD_REQUEST)
        
        config_id = AlgorithmConfig.create(data)
        config = AlgorithmConfig.find_by_id(config_id)

        # Log config creation
        log_action(
            request,
            'CREATE_CONFIG',
            company_id=cid,
            resource_type='config',
            resource_id=str(config_id),
            details={
                'name': data.get('name'),
                'algorithm': data.get('algorithm'),
            },
        )

        return Response(config, status=status.HTTP_201_CREATED)
    
    def put(self, request, pk):
        company_id = derive_company_id(request)
        config = AlgorithmConfig.find_by_id(pk)
        if not config or str(config.get('company_id')) != str(company_id):
            return Response({'error': 'Config not found'}, status=status.HTTP_404_NOT_FOUND)
        if AlgorithmConfig.update(pk, request.data):
            updated = AlgorithmConfig.find_by_id(pk)
            log_action(
                request,
                'UPDATE_CONFIG',
                company_id=company_id,
                resource_type='config',
                resource_id=str(pk),
                details={'changes': request.data},
            )
            return Response(updated)
        return Response({'error': 'Config not found'}, status=status.HTTP_404_NOT_FOUND)
    
    def delete(self, request, pk):
        company_id = derive_company_id(request)
        config = AlgorithmConfig.find_by_id(pk)
        if not config or str(config.get('company_id')) != str(company_id):
            return Response({'error': 'Config not found'}, status=status.HTTP_404_NOT_FOUND)
        if AlgorithmConfig.delete(pk):
            log_action(
                request,
                'DELETE_CONFIG',
                company_id=company_id,
                resource_type='config',
                resource_id=str(pk),
            )
            return Response(status=status.HTTP_204_NO_CONTENT)
        return Response({'error': 'Config not found'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
def activate_algorithm_config(request, pk: str):
    """Set one algorithm config as active for its company.

    Simple strategy: mark this config with field `is_active=True` and,
    if company_id is present, set others with same company_id to False.
    """
    try:
        config = AlgorithmConfig.find_by_id(pk)
        if not config:
            return Response({'error': 'Config not found'}, status=status.HTTP_404_NOT_FOUND)

        company_id = derive_company_id(request)
        if not company_id or str(config.get('company_id')) != str(company_id):
            return Response({'error': 'Config not found'}, status=status.HTTP_404_NOT_FOUND)

        # Deactivate other configs of same company (best effort)
        if company_id:
            coll = AlgorithmConfig.get_collection()
            coll.update_many({'company_id': company_id}, {'$set': {'is_active': False}})

        AlgorithmConfig.update(pk, {'is_active': True})
        updated = AlgorithmConfig.find_by_id(pk)

        # Log activation
        log_action(
            request,
            'ACTIVATE_CONFIG',
            company_id=company_id,
            resource_type='config',
            resource_id=str(pk),
            details={
                'name': updated.get('name'),
                'algorithm': updated.get('algorithm'),
            },
        )

        return Response(updated)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_schedule_conflicts(request):
    """Get schedule conflicts scoped by company."""
    try:
        company_id = derive_company_id(request)
        # Use model helper to fetch schedules filtered by company
        filt = {'company_id': company_id} if company_id else None
        schedules = Schedule.find_all(filt)
        conflicts = []
        
        # Check for time conflicts
        for i, schedule1 in enumerate(schedules):
            for schedule2 in schedules[i+1:]:
                # Same interviewer, overlapping time
                if (schedule1.get('interviewer_id') == schedule2.get('interviewer_id') and
                    schedule1.get('start_time') == schedule2.get('start_time')):
                    conflicts.append({
                        'type': 'interviewer_conflict',
                        'schedule1_id': str(schedule1.get('_id')),
                        'schedule2_id': str(schedule2.get('_id')),
                        'interviewer_id': schedule1.get('interviewer_id'),
                        'time': schedule1.get('start_time')
                    })
                
                # Same room, overlapping time
                if (schedule1.get('room_id') == schedule2.get('room_id') and
                    schedule1.get('start_time') == schedule2.get('start_time')):
                    conflicts.append({
                        'type': 'room_conflict',
                        'schedule1_id': str(schedule1.get('_id')),
                        'schedule2_id': str(schedule2.get('_id')),
                        'room_id': schedule1.get('room_id'),
                        'time': schedule1.get('start_time')
                    })
                
                # Same applicant, overlapping time
                if (schedule1.get('applicant_id') == schedule2.get('applicant_id') and
                    schedule1.get('start_time') == schedule2.get('start_time')):
                    conflicts.append({
                        'type': 'applicant_conflict',
                        'schedule1_id': str(schedule1.get('_id')),
                        'schedule2_id': str(schedule2.get('_id')),
                        'applicant_id': schedule1.get('applicant_id'),
                        'time': schedule1.get('start_time')
                    })
        
        return Response({
            'total_conflicts': len(conflicts),
            'conflicts': conflicts
        })
        
    except Exception as e:
        print(f"Error in get_schedule_conflicts: {str(e)}")
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


class PositionAPIView(APIView):
    """API endpoint for Positions"""
    
    def get(self, request, pk=None):
        company_id = derive_company_id(request)
        if pk:
            position = Position.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
            if position:
                return Response(position)
            return Response({'error': 'Position not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get only active positions by default for user's company
        is_active = request.query_params.get('is_active', 'true').lower() == 'true'
        filter_dict = {'company_id': company_id} if company_id else {}
        if request.query_params.get('is_active'):
            filter_dict['is_active'] = is_active
        positions = Position.find_all(filter_dict)
        return Response(positions)
    
    def post(self, request):
        data = request.data.copy()
        cid = derive_company_id(request)
        if not cid:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        data['company_id'] = cid
        is_valid, error_msg = Position.validate(data)
        if not is_valid:
            return Response({'error': error_msg}, status=status.HTTP_400_BAD_REQUEST)
        
        # Set defaults
        data['is_active'] = data.get('is_active', True)
        data['created_at'] = datetime.now()
        
        position_id = Position.create(data)
        position = Position.find_by_id(position_id)
        return Response(position, status=status.HTTP_201_CREATED)
    
    def put(self, request, pk):
        company_id = derive_company_id(request)
        existing = Position.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
        if not existing:
            return Response({'error': 'Position not found'}, status=status.HTTP_404_NOT_FOUND)
        
        result = Position.update(pk, request.data)
        if result:
            position = Position.find_by_id(pk)
            return Response(position)
        return Response({'error': 'Position not found'}, status=status.HTTP_404_NOT_FOUND)
    
    def delete(self, request, pk):
        company_id = derive_company_id(request)
        existing = Position.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
        if not existing:
            return Response({'error': 'Position not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if Position.delete(pk):
            return Response(status=status.HTTP_204_NO_CONTENT)
        return Response({'error': 'Position not found'}, status=status.HTTP_404_NOT_FOUND)


class InterviewSessionAPIView(APIView):
    """API endpoint for Interview Sessions"""
    
    def get(self, request, pk=None):
        company_id = derive_company_id(request)
        if pk:
            # Ensure user can only access sessions from their company
            session = InterviewSession.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
            if session:
                return Response(session)
            return Response({'error': 'Session not found'}, status=status.HTTP_404_NOT_FOUND)
        # List all sessions for user's company only
        filt = {'company_id': company_id} if company_id else {}
        sessions = InterviewSession.find_all(filt, sort=[["created_at", -1]])
        return Response(sessions)
    
    def post(self, request):
        data = request.data.copy()
        company_id = derive_company_id(request)
        if not company_id:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Force company_id from token
        data['company_id'] = company_id
        
        is_valid, error_msg = InterviewSession.validate(data)
        if not is_valid:
            return Response({'error': error_msg}, status=status.HTTP_400_BAD_REQUEST)
        
        # Set defaults
        data['is_active'] = data.get('is_active', False)
        data['created_at'] = datetime.now()
        
        session_id = InterviewSession.create(data)
        session = InterviewSession.find_by_id(session_id)
        return Response(session, status=status.HTTP_201_CREATED)
    
    def put(self, request, pk):
        company_id = derive_company_id(request)
        # Ensure session belongs to user's company
        existing = InterviewSession.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
        if not existing:
            return Response({'error': 'Session not found'}, status=status.HTTP_404_NOT_FOUND)
        
        result = InterviewSession.update(pk, request.data)
        if result:
            session = InterviewSession.find_by_id(pk)
            return Response(session)
        return Response({'error': 'Session not found'}, status=status.HTTP_404_NOT_FOUND)
    
    def delete(self, request, pk):
        company_id = derive_company_id(request)
        # Ensure session belongs to user's company
        existing = InterviewSession.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
        if not existing:
            return Response({'error': 'Session not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if InterviewSession.delete(pk):
            return Response(status=status.HTTP_204_NO_CONTENT)
        return Response({'error': 'Session not found'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
def get_active_session(request):
    """Get the currently active interview session"""
    try:
        company_id = derive_company_id(request)
        session = _get_active_session_or_none(company_id)
        if session:
            return Response(session)
        # Return 200 with empty object instead of 404 so
        # frontend callers can treat "no active" as a
        # normal state without logging errors.
        return Response({})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def set_active_session(request, pk):
    """Set a session as active. Do not automatically deactivate other sessions.

    Historically this endpoint deactivated all other sessions for the company
    which made it impossible to have multiple sessions active (useful for
    running multiple simulations, preserving historical active flags, or
    soft-activations). Change behaviour to only flip the requested session
    to `is_active=True` and leave others unchanged.
    """
    try:
        company_id = derive_company_id(request)
        if not company_id:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Ensure session belongs to user's company
        session = InterviewSession.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
        if not session:
            return Response({'error': 'Session not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Activate the specified session only (do not touch other sessions)
        result = InterviewSession.update(pk, {'is_active': True})
        if result:
            session = InterviewSession.find_by_id(pk)
            return Response(session)
        return Response({'error': 'Session not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def update_session_membership(request, pk):
    """Attach/detach applicants, interviewers, rooms, and positions to a session.
    Payload example:
    {
        "add": {"applicants": [id...], "interviewers": [], "rooms": [], "positions": []},
        "remove": {"applicants": [], "interviewers": [], "rooms": [], "positions": []}
    }
    """
    try:
        # Ensure session belongs to caller's company
        company_id = derive_company_id(request)
        if not company_id:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        session = InterviewSession.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
        if not session:
            return Response({'error': 'Session not found'}, status=status.HTTP_404_NOT_FOUND)

        payload = request.data or {}
        add = payload.get('add', {})
        remove = payload.get('remove', {})

        ops = {}
        # Build addToSet operations
        add_set = {}
        if add.get('applicants'):
            add_set['applicant_ids'] = {'$each': add['applicants']}
        if add.get('interviewers'):
            add_set['interviewer_ids'] = {'$each': add['interviewers']}
        if add.get('rooms'):
            add_set['room_ids'] = {'$each': add['rooms']}
        if add.get('positions'):
            add_set['position_ids'] = {'$each': add['positions']}
        if add_set:
            ops['$addToSet'] = add_set

        # Build pull operations
        pull = {}
        if remove.get('applicants'):
            pull['applicant_ids'] = {'$in': remove['applicants']}
        if remove.get('interviewers'):
            pull['interviewer_ids'] = {'$in': remove['interviewers']}
        if remove.get('rooms'):
            pull['room_ids'] = {'$in': remove['rooms']}
        if remove.get('positions'):
            pull['position_ids'] = {'$in': remove['positions']}
        if pull:
            ops['$pull'] = pull

        if not ops:
            return Response({'message': 'No changes requested'}, status=status.HTTP_200_OK)

        # Ensure we filter by proper ObjectId when pk is a string
        oid = ObjectId(pk) if isinstance(pk, str) else pk
        InterviewSession.get_collection().update_one({'_id': oid, 'company_id': company_id}, ops)
        session = InterviewSession.find_by_id(pk)
        if not session:
            return Response({'error': 'Session not found'}, status=status.HTTP_404_NOT_FOUND)
        return Response(session)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ApplicantAPIView(APIView):
    """API endpoint for Applicants"""

    def get(self, request, pk=None):
        company_id = derive_company_id(request)
        if pk:
            # Ensure item belongs to user's company
            item = Applicant.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
            if item:
                return Response(item)
            return Response({'error': 'Applicant not found'}, status=status.HTTP_404_NOT_FOUND)
        # Optional filters
        position = request.query_params.get('position')
        session_id = request.query_params.get('session_id')
        if session_id:
            items = InterviewSession.get_session_applicants(session_id)
            # Double-check company scoping
            if company_id:
                items = [i for i in items if i.get('company_id') == company_id]
            if position:
                items = [i for i in items if i.get('position') == position]
        else:
            filter_dict = {'company_id': company_id} if company_id else {}
            if position:
                filter_dict['position'] = position
            items = Applicant.find_all(filter_dict)
        return Response(items)

    def post(self, request):
        data = request.data.copy()
        cid = derive_company_id(request)
        if not cid:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        # Force company_id from token
        data['company_id'] = cid
        is_valid, err = Applicant.validate(data)
        if not is_valid:
            return Response({'error': err}, status=status.HTTP_400_BAD_REQUEST)
        item_id = Applicant.create(data)
        item = Applicant.find_by_id(item_id)
        return Response(item, status=status.HTTP_201_CREATED)

    def put(self, request, pk):
        company_id = derive_company_id(request)
        # Ensure item belongs to user's company
        existing = Applicant.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
        if not existing:
            return Response({'error': 'Applicant not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if Applicant.update(pk, request.data):
            return Response(Applicant.find_by_id(pk))
        return Response({'error': 'Applicant not found'}, status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, pk):
        company_id = derive_company_id(request)
        # Ensure item belongs to user's company
        existing = Applicant.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
        if not existing:
            return Response({'error': 'Applicant not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if Applicant.delete(pk):
            return Response(status=status.HTTP_204_NO_CONTENT)
        return Response({'error': 'Applicant not found'}, status=status.HTTP_404_NOT_FOUND)


class InterviewerAPIView(APIView):
    """API endpoint for Interviewers"""

    def get(self, request, pk=None):
        company_id = derive_company_id(request)
        if pk:
            item = Interviewer.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
            if item:
                return Response(item)
            return Response({'error': 'Interviewer not found'}, status=status.HTTP_404_NOT_FOUND)
        position = request.query_params.get('position')
        session_id = request.query_params.get('session_id')
        if session_id:
            items = InterviewSession.get_session_interviewers(session_id)
            # Double-check company scoping
            if company_id:
                items = [i for i in items if i.get('company_id') == company_id]
            if position:
                items = [i for i in items if i.get('position') == position]
        else:
            filter_dict = {'company_id': company_id} if company_id else {}
            if position:
                filter_dict['position'] = position
            items = Interviewer.find_all(filter_dict)
        return Response(items)

    def post(self, request):
        data = request.data.copy()
        cid = derive_company_id(request)
        if not cid:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        data['company_id'] = cid
        is_valid, err = Interviewer.validate(data)
        if not is_valid:
            return Response({'error': err}, status=status.HTTP_400_BAD_REQUEST)
        item_id = Interviewer.create(data)
        item = Interviewer.find_by_id(item_id)
        return Response(item, status=status.HTTP_201_CREATED)

    def put(self, request, pk):
        company_id = derive_company_id(request)
        existing = Interviewer.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
        if not existing:
            return Response({'error': 'Interviewer not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if Interviewer.update(pk, request.data):
            return Response(Interviewer.find_by_id(pk))
        return Response({'error': 'Interviewer not found'}, status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, pk):
        company_id = derive_company_id(request)
        existing = Interviewer.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
        if not existing:
            return Response({'error': 'Interviewer not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if Interviewer.delete(pk):
            return Response(status=status.HTTP_204_NO_CONTENT)
        return Response({'error': 'Interviewer not found'}, status=status.HTTP_404_NOT_FOUND)


class RoomAPIView(APIView):
    """API endpoint for Rooms"""

    def get(self, request, pk=None):
        company_id = derive_company_id(request)
        if pk:
            item = Room.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
            if item:
                return Response(item)
            return Response({'error': 'Room not found'}, status=status.HTTP_404_NOT_FOUND)
        session_id = request.query_params.get('session_id')
        if session_id:
            items = InterviewSession.get_session_rooms(session_id)
            # Ensure company scoping
            if company_id:
                items = [i for i in items if i.get('company_id') == company_id]
        else:
            filt = {'company_id': company_id} if company_id else {}
            items = Room.find_all(filt)
        return Response(items)

    def post(self, request):
        data = request.data.copy()
        cid = derive_company_id(request)
        if not cid:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        data['company_id'] = cid
        is_valid, err = Room.validate(data)
        if not is_valid:
            return Response({'error': err}, status=status.HTTP_400_BAD_REQUEST)
        item_id = Room.create(data)
        item = Room.find_by_id(item_id)
        return Response(item, status=status.HTTP_201_CREATED)

    def put(self, request, pk):
        company_id = derive_company_id(request)
        existing = Room.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
        if not existing:
            return Response({'error': 'Room not found'}, status=status.HTTP_404_NOT_FOUND)
        data = request.data.copy()
        data['company_id'] = company_id
        if Room.update(pk, data):
            return Response(Room.find_by_id(pk))
        return Response({'error': 'Room not found'}, status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, pk):
        company_id = derive_company_id(request)
        existing = Room.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
        if not existing:
            return Response({'error': 'Room not found'}, status=status.HTTP_404_NOT_FOUND)
        if Room.delete(pk):
            return Response(status=status.HTTP_204_NO_CONTENT)
        return Response({'error': 'Room not found'}, status=status.HTTP_404_NOT_FOUND)


class ScheduleAPIView(APIView):
    """API endpoint for Schedules"""

    def get(self, request, pk=None):
        company_id = derive_company_id(request)
        if pk:
            item = Schedule.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
            if item:
                return Response(item)
            return Response({'error': 'Schedule not found'}, status=status.HTTP_404_NOT_FOUND)
        session_id = request.query_params.get('session_id')
        filter_dict = {'company_id': company_id} if company_id else {}
        if session_id:
            filter_dict['session_id'] = session_id
        items = Schedule.find_all(filter_dict)
        return Response(items)

    def post(self, request):
        data = request.data.copy()
        cid = derive_company_id(request)
        if not cid:
            return Response({'error': 'Unable to resolve company from token'}, status=status.HTTP_400_BAD_REQUEST)
        data['company_id'] = cid
        is_valid, err = Schedule.validate(data)
        if not is_valid:
            return Response({'error': err}, status=status.HTTP_400_BAD_REQUEST)
        item_id = Schedule.create(data)
        item = Schedule.find_by_id(item_id)
        return Response(item, status=status.HTTP_201_CREATED)

    def put(self, request, pk):
        company_id = derive_company_id(request)
        existing = Schedule.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
        if not existing:
            return Response({'error': 'Schedule not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if Schedule.update(pk, request.data):
            return Response(Schedule.find_by_id(pk))
        return Response({'error': 'Schedule not found'}, status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, pk):
        company_id = derive_company_id(request)
        existing = Schedule.find_one({'_id': ObjectId(str(pk)), 'company_id': company_id})
        if not existing:
            return Response({'error': 'Schedule not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if Schedule.delete(pk):
            return Response(status=status.HTTP_204_NO_CONTENT)
        return Response({'error': 'Schedule not found'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
def get_schedule_timeline(request):
    """Return schedules grouped by room for timeline view, scoped to company."""
    try:
        session_id = request.query_params.get('session_id')
        company_id = derive_company_id(request)
        filter_dict = {'company_id': company_id}
        if session_id:
            # Validate session belongs to company
            session = InterviewSession.find_one({'_id': ObjectId(str(session_id)), 'company_id': company_id})
            if not session:
                return Response({'error': 'Session not found'}, status=status.HTTP_404_NOT_FOUND)
            filter_dict['session_id'] = session_id
        schedules = Schedule.find_all(filter_dict)

        # Build lookup maps with stringified ObjectIds for reliable matching
        # Prefer session-scoped entities when session_id is provided, otherwise company-scoped
        if session_id:
            apps_list = InterviewSession.get_session_applicants(session_id)
            ints_list = InterviewSession.get_session_interviewers(session_id)
            rooms_list = InterviewSession.get_session_rooms(session_id)
        else:
            apps_list = Applicant.find_all({'company_id': company_id} if company_id else None)
            ints_list = Interviewer.find_all({'company_id': company_id} if company_id else None)
            rooms_list = Room.find_all({'company_id': company_id} if company_id else None)
        applicants = {str(a.get('_id')): a for a in apps_list}
        interviewers = {str(i.get('_id')): i for i in ints_list}
        rooms = {str(r.get('_id')): r for r in rooms_list}

        timeline_grouped = {}
        for s in schedules:
            rid_raw = s.get('room_id')
            rid = str(rid_raw)
            room = rooms.get(rid) or rooms.get(str(rid_raw)) or {}
            # Prefer code then name; fallback to shortened id for aesthetics
            room_key = room.get('room_code') or room.get('room_name') or rid[:8]
            if room_key not in timeline_grouped:
                timeline_grouped[room_key] = []

            aid_raw = s.get('applicant_id')
            iid_raw = s.get('interviewer_id')
            aid = str(aid_raw)
            iid = str(iid_raw)
            applicant = applicants.get(aid) or {}
            interviewer = interviewers.get(iid) or {}

            timeline_grouped[room_key].append({
                'id': s.get('_id'),
                'applicant': applicant.get('full_name', 'Applicant'),
                'interviewer': interviewer.get('full_name', 'Interviewer'),
                'position': s.get('position') or applicant.get('position'),
                'start': s.get('start_time'),
                'end': s.get('end_time'),
                'interview_date': s.get('interview_date'),
                'status': s.get('status', 'scheduled'),
            })

        return Response(timeline_grouped)
    except Exception as e:
        print(f"Error in get_schedule_timeline: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ---------------------------------------------------------------------------
# GA Dev / Debug endpoints (stub implementation for frontend Dev Mode)
# ---------------------------------------------------------------------------


@api_view(['GET'])
def latest_debug_run(request):
    """Return a minimal stub GA debug run so Dev Mode can render.

    This does NOT run a full GA; it just inspects recent ScheduleResult
    documents for the caller's company and builds a lightweight structure
    with a single generation and a few representative individuals.
    """
    try:
        company_id = derive_company_id(request)
        if not company_id:
            return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)

        # Use last few ScheduleResult docs as pseudo "representative individuals"
        results = ScheduleResult.find_all(
            {'company_id': company_id},
            limit=5,
            sort=[('created_at', -1)],
        )
        reps = []
        for idx, r in enumerate(results, start=1):
            reps.append({
                'individual_id': str(r.get('_id')),
                'rank': idx,
                'fitness': r.get('fitness_score', 0.0),
                'selection_meta': {
                    'is_selected_parent': False,
                    'selection_prob': None,
                },
                'schedule_snapshot': [
                    {
                        'applicant_name': s.get('applicant_id'),
                        'interviewer_name': s.get('interviewer_id'),
                        'room_name': s.get('room_id'),
                        'position': s.get('position'),
                        'start_time': s.get('start_time'),
                        'end_time': s.get('end_time'),
                        'date': s.get('interview_date'),
                        'has_conflict': False,
                    }
                    for s in r.get('schedule_data', [])
                ],
            })

        payload = {
            'run_id': 'stub',
            'generations': [
                {
                    'index': 0,
                    'stats': {
                        'best_fitness': max((r.get('fitness', 0.0) for r in reps), default=0.0),
                    },
                    'representatives': reps,
                }
            ],
        }
        return Response(payload)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def debug_schedule_detail(request):
    """Stub schedule detail for mutation view.

    Currently just returns the same snapshot for before/after mutation and
    an empty mutation list, so frontend Dev Mode can render without errors.
    """
    try:
        company_id = derive_company_id(request)
        if not company_id:
            return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)

        individual_id = request.query_params.get('individual_id')
        if not individual_id:
            return Response({'error': 'individual_id required'}, status=status.HTTP_400_BAD_REQUEST)

        res = ScheduleResult.find_one({'_id': ObjectId(str(individual_id)), 'company_id': company_id})
        if not res:
            return Response({'error': 'ScheduleResult not found'}, status=status.HTTP_404_NOT_FOUND)

        base = {
            'individual_id': str(res.get('_id')),
            'rank': 1,
            'fitness': res.get('fitness_score', 0.0),
            'schedule_snapshot': [
                {
                    'applicant_name': s.get('applicant_id'),
                    'interviewer_name': s.get('interviewer_id'),
                    'room_name': s.get('room_id'),
                    'position': s.get('position'),
                    'start_time': s.get('start_time'),
                    'end_time': s.get('end_time'),
                    'date': s.get('interview_date'),
                }
                for s in res.get('schedule_data', [])
            ],
        }

        return Response({
            'before_mutation': base,
            'after_mutation': base,
            'mutations': [],
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def debug_crossover_detail(request):
    """Stub crossover detail: returns the same individual as both parents and child."""
    try:
        company_id = derive_company_id(request)
        if not company_id:
            return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)

        child_id = request.query_params.get('child_id')
        if not child_id:
            return Response({'error': 'child_id required'}, status=status.HTTP_400_BAD_REQUEST)

        res = ScheduleResult.find_one({'_id': ObjectId(str(child_id)), 'company_id': company_id})
        if not res:
            return Response({'error': 'ScheduleResult not found'}, status=status.HTTP_404_NOT_FOUND)

        def to_ind(doc):
            return {
                'individual_id': str(doc.get('_id')),
                'rank': 1,
                'fitness': doc.get('fitness_score', 0.0),
                'schedule_snapshot': [
                    {
                        'applicant_name': s.get('applicant_id'),
                        'interviewer_name': s.get('interviewer_id'),
                        'room_name': s.get('room_id'),
                        'position': s.get('position'),
                        'start_time': s.get('start_time'),
                        'end_time': s.get('end_time'),
                        'date': s.get('interview_date'),
                    }
                    for s in doc.get('schedule_data', [])
                ],
            }

        ind = to_ind(res)
        return Response({
            'parent_a': ind,
            'parent_b': ind,
            'child_before': ind,
            'child_after': ind,
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def debug_lineage_detail(request):
    """Stub lineage detail: returns the selected individual as the only parent and child."""
    try:
        company_id = derive_company_id(request)
        if not company_id:
            return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)

        individual_id = request.query_params.get('individual_id')
        if not individual_id:
            return Response({'error': 'individual_id required'}, status=status.HTTP_400_BAD_REQUEST)

        res = ScheduleResult.find_one({'_id': ObjectId(str(individual_id)), 'company_id': company_id})
        if not res:
            return Response({'error': 'ScheduleResult not found'}, status=status.HTTP_404_NOT_FOUND)

        def to_ind(doc):
            return {
                'individual_id': str(doc.get('_id')),
                'rank': 1,
                'fitness': doc.get('fitness_score', 0.0),
                'schedule_snapshot': [
                    {
                        'applicant_name': s.get('applicant_id'),
                        'interviewer_name': s.get('interviewer_id'),
                        'room_name': s.get('room_id'),
                        'position': s.get('position'),
                        'start_time': s.get('start_time'),
                        'end_time': s.get('end_time'),
                        'date': s.get('interview_date'),
                    }
                    for s in doc.get('schedule_data', [])
                ],
            }

        ind = to_ind(res)
        return Response({'parents': [ind], 'child': ind})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

